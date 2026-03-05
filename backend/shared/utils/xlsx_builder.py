from __future__ import annotations
from io import BytesIO
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

RAV_HEADERS = ["Datum","Firma","Stellenbezeichnung","Ort","Kanal","Status","Nachweis (Pfad/Link)","Notizen"]

def build_rav_xlsx(rows: List[Dict]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Arbeitsbemühungen"
    ws.append(RAV_HEADERS)
    for r in rows:
        ws.append([r.get("date",""), r.get("company",""), r.get("title",""), r.get("location",""), r.get("channel",""), r.get("status",""), r.get("proof",""), r.get("notes","")])
    for col in range(1, len(RAV_HEADERS)+1):
        max_len = 12
        for cell in ws[get_column_letter(col)]:
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)
    bio = BytesIO(); wb.save(bio); return bio.getvalue()
